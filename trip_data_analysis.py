import time

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# 
df = pd.read_excel("trip.xlsx")


def shift_category(row):
    if row['Shift type'] == "night":
        return 50
    else:
        return 0
    


df['extra cost'] = df.apply(shift_category, axis=1)
print(df)

# length of name column
# df['name_length'] = df['name'].apply(len)


# convert trip ID to string
df['trip ID'] = df['trip ID'].astype(str)

# Enter the date you want to filter by
search_date = "2026-06-05"

# Filter employees by date
employee = df[df['Date'] == search_date]

# print the names of the employees who took a trip on the specified date
print(employee['name'])


# dates on only female employees took trips
female_trips = df[df['gender'] == 'female']

# print the names of the female employees who took trips
print(female_trips['Date'])

# Sum of Per Head Cost column
total_cost = df['per head cost'].sum()

print (f"Total cost of all trips: {total_cost}")

# open workbook and select sheet
wb = load_workbook("trip.xlsx")
ws = wb.active

# Create thick border
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

ws['G8'] = total_cost
ws["G8"].alignment = Alignment(horizontal='center', vertical='center')
ws["G8"].border = thick_border

print(wb.sheetnames)
# Save file
wb.save("trip.xlsx")

df.to_excel("employees_updated.xlsx", index=False)




